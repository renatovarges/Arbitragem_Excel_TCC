import streamlit as st
import traceback
import io

st.set_page_config(
    page_title="Arbitragem — Planilha Excel",
    page_icon="📊",
    layout="centered",
)

st.title("📊 Análise de Arbitragem — Planilha Excel")
st.caption("Escolha a rodada e baixe os dados completos em Excel.")

try:
    import json
    from pathlib import Path
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    import arbitragem_scraper as scraper

    # ── Rodadas disponíveis ──────────────────────────────────────────────────
    with open(Path(__file__).parent / "rodadas.json", encoding="utf-8") as f:
        rodadas_data = json.load(f)
    rodadas_disponiveis = sorted(int(k) for k in rodadas_data.keys())

    # ── Sidebar ─────────────────────────────────────────────────────────────
    st.sidebar.header("⚙️ Configurações")
    rodada_sel = st.sidebar.selectbox(
        "Rodada",
        rodadas_disponiveis,
        index=min(17, len(rodadas_disponiveis) - 1),
        format_func=lambda x: f"Rodada {x}",
    )

    with st.sidebar.expander("ℹ️ Sobre a planilha"):
        st.markdown("""
        A planilha Excel gerada contém:
        - **Aba principal**: todos os jogos da rodada com árbitro e estatísticas
        - Formatação com cores por coluna
        - Pronto para importar ou compartilhar

        > Se a escala ainda não foi divulgada, o árbitro aparece como **"A confirmar"**.
        """)

    # ── Preview dos confrontos ──────────────────────────────────────────────
    confrontos_base = rodadas_data.get(str(rodada_sel), [])
    st.subheader(f"📅 Confrontos — Rodada {rodada_sel}")
    cols = st.columns(2)
    for i, c in enumerate(confrontos_base):
        cols[i % 2].markdown(f"**{c['mandante']}** × {c['visitante']}")

    st.divider()

    # ── Geração ─────────────────────────────────────────────────────────────
    if st.button("📊 GERAR PLANILHA EXCEL", use_container_width=True, type="primary"):
        with st.spinner("Buscando árbitros e estatísticas..."):
            try:
                dados = scraper.get_rodada_completa(rodada_sel)
            except Exception as e:
                st.error(f"Erro ao buscar dados: {e}")
                st.code(traceback.format_exc())
                st.stop()

        # Preview dos dados
        st.subheader("📊 Dados coletados")
        df_preview = pd.DataFrame([{
            "Mandante":  d["mandante"],
            "Visitante": d["visitante"],
            "Árbitro":   d["arbitro"],
            "Jogos":     d["jogos"],
            "Pênaltis":  d["penaltis"],
            "Faltas (Média)":   d["faltas_media"],
            "Amarelos (Média)": d["amarelos_media"],
            "Vermelhos (Total)":d["vermelhos_total"],
        } for d in dados])
        st.dataframe(df_preview, use_container_width=True)

        # Verificar escala
        confirmados = [d for d in dados if d["arbitro"] != "A confirmar"]
        if not confirmados:
            st.warning(
                "⚠️ A escala desta rodada ainda não foi divulgada. "
                "A planilha será gerada com 'A confirmar' nos campos de árbitro."
            )

        # Gerar Excel
        with st.spinner("Montando planilha Excel..."):
            try:
                excel_bytes = _build_excel(dados, rodada_sel)
                st.success("Planilha gerada! ✅")
                st.download_button(
                    label="⬇️ BAIXAR PLANILHA EXCEL",
                    data=excel_bytes,
                    file_name=f"Arbitragem_Rodada_{rodada_sel}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Erro ao gerar Excel: {e}")
                st.code(traceback.format_exc())

except Exception as e:
    st.error(f"⚠️ Erro inesperado: {e}")
    st.code(traceback.format_exc())


# ── FUNÇÃO DE GERAÇÃO DO EXCEL ───────────────────────────────────────────────

def _build_excel(dados: list, rodada_num: int) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    wb = Workbook()
    ws = wb.active
    ws.title = f"Rodada {rodada_num}"

    # ── Estilo base ──────────────────────────────────────────────────────────
    GREEN_DARK  = "1A4D2E"
    GREEN_LIGHT = "C6EFCE"
    YELLOW_L    = "FFEB9C"
    RED_L       = "FFC7CE"
    HEADER_BG   = "1A4D2E"
    HEADER_FG   = "FFFFFF"
    ALT_ROW     = "F2F2F2"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold_white = Font(bold=True, color=HEADER_FG, size=11)
    bold_dark  = Font(bold=True, color="111111", size=11)

    def fill(hex_color: str):
        return PatternFill("solid", fgColor=hex_color)

    def center(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    # ── Linha de título (merged) ──────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws["A1"] = f"ANÁLISE DE ARBITRAGEM — RODADA {rodada_num}"
    ws["A1"].font      = Font(bold=True, color=HEADER_FG, size=14)
    ws["A1"].fill      = fill(HEADER_BG)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    # ── Linha de subtítulo ────────────────────────────────────────────────────
    ws.merge_cells("A2:H2")
    ws["A2"] = "Estatísticas da temporada 2026  ·  Fonte: Transfermarkt / Sofascore"
    ws["A2"].font      = Font(italic=True, color="666666", size=9)
    ws["A2"].fill      = fill("EEEEEE")
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 16

    # ── Cabeçalhos das colunas ────────────────────────────────────────────────
    headers = [
        "MANDANTE", "VISITANTE", "ÁRBITRO",
        "JOGOS", "PÊNALTIS", "FALTAS (Média)",
        "AMARELOS (Média)", "VERMELHOS (Total)",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font      = bold_white
        cell.fill      = fill(HEADER_BG)
        cell.alignment = center(wrap=True)
        cell.border    = border
    ws.row_dimensions[3].height = 22

    # ── Dados ────────────────────────────────────────────────────────────────
    for r_idx, d in enumerate(dados, start=4):
        row_fill = fill("FFFFFF") if r_idx % 2 == 0 else fill(ALT_ROW)
        values = [
            d["mandante"],
            d["visitante"],
            d["arbitro"],
            d["jogos"],
            d["penaltis"],
            d["faltas_media"],
            d["amarelos_media"],
            d["vermelhos_total"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val if val != "—" else None)
            cell.font      = Font(size=11, bold=(col_idx == 3))
            cell.fill      = row_fill
            cell.alignment = center()
            cell.border    = border
        ws.row_dimensions[r_idx].height = 20

    # ── Larguras das colunas ─────────────────────────────────────────────────
    col_widths = [18, 18, 28, 8, 10, 14, 16, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Formatação condicional (escala de cores) ──────────────────────────────
    # Pênaltis (col E): verde para alto
    ws.conditional_formatting.add(
        f"E4:E{3+len(dados)}",
        ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            end_type="max",   end_color="63BE7B",
        )
    )
    # Faltas (col F): vermelho para alto
    ws.conditional_formatting.add(
        f"F4:F{3+len(dados)}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max",   end_color="F8696B",
        )
    )
    # Amarelos (col G): vermelho para alto
    ws.conditional_formatting.add(
        f"G4:G{3+len(dados)}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max",   end_color="F8696B",
        )
    )
    # Vermelhos (col H): vermelho para alto
    ws.conditional_formatting.add(
        f"H4:H{3+len(dados)}",
        ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            end_type="max",   end_color="F8696B",
        )
    )

    # ── Congelar cabeçalhos ───────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Salvar em memória ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
