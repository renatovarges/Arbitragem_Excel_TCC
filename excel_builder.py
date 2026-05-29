"""
excel_builder.py — monta a planilha Excel formatada da análise de arbitragem.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


def build_excel(dados: list, rodada_num: int) -> bytes:
    """Gera os bytes de um .xlsx formatado para a rodada."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Rodada {rodada_num}"

    HEADER_BG = "1A4D2E"
    HEADER_FG = "FFFFFF"
    ALT_ROW   = "F2F2F2"

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def center(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    n = len(dados)

    # ── Título ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws["A1"] = f"ANÁLISE DE ARBITRAGEM — RODADA {rodada_num}"
    ws["A1"].font      = Font(bold=True, color=HEADER_FG, size=14)
    ws["A1"].fill      = fill(HEADER_BG)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    # ── Subtítulo ─────────────────────────────────────────────────────────
    ws.merge_cells("A2:H2")
    ws["A2"] = "Estatísticas da temporada 2026  ·  Fonte: Transfermarkt / Sofascore"
    ws["A2"].font      = Font(italic=True, color="666666", size=9)
    ws["A2"].fill      = fill("EEEEEE")
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 16

    # ── Cabeçalhos ────────────────────────────────────────────────────────
    headers = [
        "MANDANTE", "VISITANTE", "ÁRBITRO",
        "JOGOS", "PÊNALTIS", "FALTAS (Média)",
        "AMARELOS (Média)", "VERMELHOS (Total)",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font      = Font(bold=True, color=HEADER_FG, size=11)
        cell.fill      = fill(HEADER_BG)
        cell.alignment = center(wrap=True)
        cell.border    = border
    ws.row_dimensions[3].height = 22

    # ── Dados ─────────────────────────────────────────────────────────────
    for r_idx, d in enumerate(dados, start=4):
        row_fill = fill("FFFFFF") if r_idx % 2 == 0 else fill(ALT_ROW)
        values = [
            d["mandante"], d["visitante"], d["arbitro"],
            d["jogos"], d["penaltis"], d["faltas_media"],
            d["amarelos_media"], d["vermelhos_total"],
        ]
        for col_idx, val in enumerate(values, start=1):
            v = val if val not in ("—", "", None) else None
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.font      = Font(size=11, bold=(col_idx == 3))
            cell.fill      = row_fill
            cell.alignment = center()
            cell.border    = border
        ws.row_dimensions[r_idx].height = 20

    # ── Larguras ──────────────────────────────────────────────────────────
    for i, w in enumerate([18, 18, 28, 8, 10, 14, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last = 3 + n

    # ── Formatação condicional (escala de cores) ──────────────────────────
    # Pênaltis (E): mais = verde
    ws.conditional_formatting.add(
        f"E4:E{last}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max", end_color="63BE7B"),
    )
    # Faltas (F): mais = vermelho
    ws.conditional_formatting.add(
        f"F4:F{last}",
        ColorScaleRule(start_type="min", start_color="63BE7B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="F8696B"),
    )
    # Amarelos (G): mais = vermelho
    ws.conditional_formatting.add(
        f"G4:G{last}",
        ColorScaleRule(start_type="min", start_color="63BE7B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="F8696B"),
    )
    # Vermelhos (H): mais = vermelho
    ws.conditional_formatting.add(
        f"H4:H{last}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max", end_color="F8696B"),
    )

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
